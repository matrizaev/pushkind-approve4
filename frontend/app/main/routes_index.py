from datetime import datetime, timezone

from flask import render_template, flash, request, redirect, url_for, Response
from flask import current_app
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from app.email import SendEmail
from app.main import bp
from app.utils import role_forbidden, role_required
from app.main.utils import send_email_notification, GetNewOrderNumber
from app.utils import get_filter_timestamps
from app.main.forms import MergeOrdersForm, SaveOrdersForm
from app.api.order import OrderApi
from app.api.project import ProjectApi
from app.api.hub import CategoryApi


################################################################################
# Index page
################################################################################

@bp.route('/')
@bp.route('/index/')
@login_required
@role_forbidden(['default'])
def show_index():

    dates = get_filter_timestamps()
    filter_from = request.args.get('from', default=dates['recently']['value'], type=int)
    filter_disapproved = request.args.get('disapproved', default=None, type=str)
    if filter_disapproved is not None:
        filter_disapproved = True

    if current_user.role.name in ['purchaser', 'validator']:
        filter_focus = request.args.get('focus', default=None, type=str)
        if filter_focus is not None:
            filter_focus = True
    else:
        filter_focus = None

    orders = OrderApi.get_entities(filters={
        'from': filter_from,
        'focus': filter_focus,
        'disapproved': filter_disapproved
    }) or []

    # if filter_disapproved is None:
    #     orders = orders.filter(
    #         ~Order.status.in_([OrderStatus.not_approved, OrderStatus.cancelled])
    #     )

    # if current_user.role == UserRoles.initiative:
    #     orders = orders.filter(Order.initiative_id == current_user.id)

    # if filter_from > 0:
    #     orders = orders.filter(Order.create_timestamp > filter_from)

    # if current_user.role in [UserRoles.purchaser, UserRoles.validator]:

    #     if filter_focus is None:
    #         if current_user.role == UserRoles.validator:
    #             orders = (
    #                 orders
    #                 .filter(Order.status != OrderStatus.approved)
    #                 .filter(~Order.user_approvals.any(OrderApproval.user_id == current_user.id))
    #             )
    #         elif current_user.role == UserRoles.purchaser:
    #             orders = orders.filter(Order.dealdone == False)
    #         orders = orders.filter(~Order.children.any())

    #     orders = orders.join(OrderCategory)
    #     orders = orders.filter(
    #         OrderCategory.category_id.in_(cat.id for cat in current_user.categories)
    #     )
    #     orders = orders.join(Site)
    #     orders = orders.filter(Site.project_id.in_(p.id for p in current_user.projects))

    # if current_user.role == UserRoles.supervisor:
    #     orders = orders.join(Site)
    #     orders = orders.join(Project).filter(Project.enabled == True)

    # if current_user.role == UserRoles.vendor:
    #     vendor = Vendor.query.filter_by(
    #         hub_id=current_user.hub_id,
    #         email=current_user.email
    #     ).first()
    #     if vendor:
    #         orders = orders.filter(
    #             # Order.purchased == True,
    #             Order.vendors.any(OrderVendor.vendor_id == vendor.id)
    #         )
    #     else:
    #         orders = orders.filter(
    #             # Order.purchased == True,
    #             Order.vendors.any(OrderVendor.vendor_id == None)
    #         )

    return render_template(
        'index.html',
        orders=orders,
        dates=dates,
        filter_from=filter_from,
        filter_focus=filter_focus,
        filter_disapproved=filter_disapproved,
        merge_form=MergeOrdersForm(),
        save_form=SaveOrdersForm(orders=[order['id'] for order in orders])
    )


@bp.route('/orders/merge/', methods=['POST'])
@login_required
@role_required(['admin', 'initiative', 'purchaser'])
def MergeOrders():
    return '<p>MergeOrders</p>'
    # form = MergeOrdersForm()
    # if form.validate_on_submit():
    #     orders_list = form.orders.data
    #     if not isinstance(orders_list, list) or len(orders_list) < 2:
    #         flash('Некорректный список заявок.')
    #         return redirect(url_for('main.show_index'))

    #     orders = []

    #     orders = Order.query.filter(Order.id.in_(orders_list), Order.hub_id == current_user.hub_id)
    #     orders = orders.filter(~Order.children.any())
    #     if current_user.role == UserRoles.initiative:
    #         orders = orders.filter(Order.initiative_id == current_user.id)

    #     orders = orders.all()

    #     if len(orders) < 2:
    #         flash('Некорректный список заявок.')
    #         return redirect(url_for('main.show_index'))

    #     for order in orders[1:]:
    #         if order.site_id != orders[0].site_id or \
    #                 order.income_statement != orders[0].income_statement or \
    #                 order.cashflow_statement != orders[0].cashflow_statement:
    #             flash('Нельзя объединять заявки с разными объектами, БДДР или БДДС.')
    #             return redirect(url_for('main.show_index'))

    #     products = {}
    #     categories = []
    #     vendors = []
    #     for order in orders:
    #         categories += [cat.id for cat in order.categories]
    #         vendors += [v.id for v in order.vendors]
    #         for product in order.products:
    #             if 'selectedOptions' in product and len(product['selectedOptions']) > 1:
    #                 product_id = product['sku'] + ''.join(sorted(
    #                     [k['value'] for k in product['selectedOptions']]
    #                 ))
    #             else:
    #                 product_id = product['sku']
    #             if product_id not in products:
    #                 products[product_id] = {}
    #                 products[product_id]['sku'] = product['sku']
    #                 products[product_id]['id'] = abs(hash(product_id))
    #                 products[product_id]['name'] = product['name']
    #                 products[product_id]['price'] = product['price']
    #                 products[product_id]['quantity'] = product['quantity']
    #                 if 'selectedOptions' in product:
    #                     products[product_id]['selectedOptions'] = product['selectedOptions']
    #                 products[product_id]['categoryId'] = product['categoryId']
    #                 products[product_id]['imageUrl'] = product['imageUrl']
    #                 if 'vendor' in product:
    #                     products[product_id]['vendor'] = product['vendor']
    #                 if 'category' in product:
    #                     products[product_id]['category'] = product['category']
    #             else:
    #                 products[product_id]['quantity'] += product['quantity']

    #     order_number = GetNewOrderNumber()
    #     order = Order(number = order_number)
    #     db.session.add(order)
    #     order.initiative = current_user

    #     now = datetime.now(tz=timezone.utc)

    #     order.products = [product for _,product in products.items()]
    #     order.total = sum([product['quantity']*product['price'] for product in order.products])
    #     order.income_id = orders[0].income_id
    #     order.cashflow_id = orders[0].cashflow_id
    #     order.site_id = orders[0].site_id
    #     order.status = OrderStatus.new
    #     order.create_timestamp = int(now.timestamp())

    #     order.hub_id = current_user.hub_id
    #     order.categories = Category.query.filter(
    #         Category.id.in_(categories),
    #         Category.hub_id == current_user.hub_id
    #     ).all()
    #     order.vendors = Vendor.query.filter(
    #         Vendor.name.in_(vendors),
    #         Vendor.hub_id == current_user.hub_id
    #     ).all()
    #     order.parents = orders

    #     message = 'заявка объединена из заявок'

    #     for o in orders:
    #         o.total = 0.0
    #         message += f' {o.number}'
    #         message2 = f'заявка объединена в заявку {order.number}'
    #         event = OrderEvent(
    #             user_id=current_user.id,
    #             order_id=o.id,
    #             type=EventType.merged,
    #             data=message2,
    #             timestamp=datetime.now(tz=timezone.utc)
    #         )
    #         db.session.add(event)

    #     event = OrderEvent(
    #         user_id=current_user.id,
    #         order_id=order.id,
    #         type=EventType.merged,
    #         data=message,
    #         timestamp=datetime.now(tz=timezone.utc)
    #     )
    #     db.session.add(event)

    #     db.session.commit()

    #     order.update_positions()

    #     flash(f'Объединено заявок: {len(orders)}. Идентификатор новой заявки {order.number}')

    #     send_email_notification('new', order)
    # else:
    #     for error in form.orders.errors:
    #         flash(error)
    # return redirect(url_for('main.show_index'))


@bp.route('/orders/save/', methods=['POST'])
@login_required
@role_forbidden(['default'])
def SaveOrders():
    return '<p>SaveOrders</p>'
    # form = SaveOrdersForm()
    # if form.validate_on_submit():
    #     orders_list = form.orders.data
    #     if not isinstance(orders_list, list):
    #         flash('Некорректный список заявок.')
    #         return redirect(url_for('main.show_index'))

    #     orders = []

    #     orders = Order.query.filter(Order.id.in_(orders_list), Order.hub_id == current_user.hub_id)
    #     if current_user.role == UserRoles.initiative:
    #         orders = orders.filter(Order.initiative_id == current_user.id)

    #     orders = orders.all()

    #     wb = Workbook()

    #     ws = wb.active

    #     ws['A1'] = 'Номер'
    #     ws['B1'] = 'Дата'
    #     ws['C1'] = 'Проект'
    #     ws['D1'] = 'Объект'
    #     ws['E1'] = 'Сумма'
    #     ws['F1'] = 'Позиций'
    #     ws['G1'] = 'Статус'
    #     ws['H1'] = 'Инициатор'
    #     ws['I1'] = 'Статья БДР'
    #     ws['J1'] = 'Статья БДДС'
    #     ws['K1'] = 'Кем согласована'
    #     ws['L1'] = 'Ждём согласования'
    #     ws['M1'] = 'Категории'

    #     for i, order in enumerate(orders, start=2):
    #         ws.cell(row=i, column=1, value=order.number)
    #         ws.cell(row=i, column=2, value=datetime.fromtimestamp(order.create_timestamp))
    #         if order.site is not None:
    #             ws.cell(row=i, column=3, value=order.site.project.name)
    #             ws.cell(row=i, column=4, value=order.site.name)
    #         ws.cell(row=i, column=5, value=order.total)
    #         ws.cell(row=i, column=6, value=len(order.products))
    #         ws.cell(row=i, column=7, value=str(order.status))
    #         ws.cell(row=i, column=8, value=order.initiative.name)
    #         ws.cell(
    #             row=i,
    #             column=9,
    #             value=order.income_statement.name if order.income_statement is not None else ''
    #         )
    #         ws.cell(
    #             row=i,
    #             column=10,
    #             value=order.cashflow_statement.name if order.cashflow_statement is not None else ''
    #         )
    #         ws.cell(
    #             row=i,
    #             column=11,
    #             value=', '.join(
    #                 pos.position.name for pos in order.approvals if pos.approved is True
    #             )
    #         )
    #         ws.cell(
    #             row=i,
    #             column=12,
    #             value=', '.join(
    #                 pos.position.name for pos in order.approvals if pos.approved is False
    #             )
    #         )
    #         ws.cell(row=i, column=13, value=', '.join([cat.name for cat in order.categories]))

    #     data = save_virtual_workbook(wb)
    #     return Response(
    #         data,
    #         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #         headers={'Content-Disposition': 'attachment;filename=export.xlsx'}
    #     )

    # for error in form.orders.errors:
    #     flash(error)
    # return redirect(url_for('main.show_index'))


@bp.route('/support/call/', methods=['POST'])
@login_required
@role_forbidden(['default'])
def call_support():
    comment = request.form.get('comment', '', type=str)
    if len(comment) > 0 and len(comment) < 2048:
        SendEmail(
            'Обращение в поддержку',
            current_app.config['MAIL_USERNAME'],
            [current_app.config['ADMIN_EMAIL']],
            text_body=render_template('email/support.txt', comment=comment),
            html_body=render_template('email/support.html', comment=comment),
        )
        flash('Сообщение отправлено в поддержку.')
    else:
        flash('Сообщение некорректной длины (максимум 2048 символов).')
    return redirect(url_for('main.show_index'))
