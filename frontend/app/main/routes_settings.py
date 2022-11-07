from flask_login import current_user, login_required
from flask import render_template, redirect, url_for, flash
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from app.main import bp
from app.main.forms import UserRolesForm, UserSettingsForm
from app.utils import role_required, role_forbidden
from app.api.project import ProjectApi
from app.api.hub import CategoryApi
from app.api.user import RoleApi, UserApi


################################################################################
# Settings page
################################################################################


@bp.route('/users/show', methods=['GET'])
@login_required
@role_forbidden(['default', 'vendor'])
def show_users():

    projects = ProjectApi.get_entities(enabled=1) or []
    categories = CategoryApi.get_entities() or []
    roles = RoleApi.get_entities() or []

    if current_user.role.name == 'admin':
        user_form = UserRolesForm()
        user_form.role.choices = [
            (r['id'], r['pretty']) for r in roles
        ]
    else:
        user_form = UserSettingsForm()

    if current_user.role.name in ('admin', 'purchaser', 'validator'):
        user_form.about_user.categories.choices = [
            (c['id'], c['name']) for c in categories
        ]
        user_form.about_user.projects.choices = [
            (p['id'], p['name']) for p in projects
        ]
    else:
        user_form.about_user.categories.choices = []
        user_form.about_user.projects.choices = []

    user = UserApi.get_entities(id=current_user.id)[0]

    if current_user.role.name == 'admin':
        users = UserApi.get_entities() or []
        return render_template('settings.html', user_form=user_form, users=users, user=user)

    return render_template('settings.html', user_form=user_form, user=user)


@bp.route('/user/edit.', methods=['POST'])
@login_required
@role_forbidden(['default', 'vendor'])
def edit_user():

    projects = ProjectApi.get_entities(enabled=1) or []
    categories = CategoryApi.get_entities() or []
    roles = RoleApi.get_entities() or []

    if current_user.role.name == 'admin':
        user_form = UserRolesForm()
        user_form.role.choices = [
            (r['id'], r['pretty']) for r in roles
        ]
    else:
        user_form = UserSettingsForm()

    if current_user.role.name in ('admin', 'purchaser', 'validator'):
        user_form.about_user.categories.choices = [
            (c['id'], c['name']) for c in categories
        ]
        user_form.about_user.projects.choices = [
            (p['id'], p['name']) for p in projects
        ]
    else:
        user_form.about_user.categories.choices = []
        user_form.about_user.projects.choices = []

    if user_form.submit.data:
        if user_form.validate_on_submit():
            if current_user.role.name == 'admin':
                user_role = next(filter(lambda x: x['id']==user_form.role.data, roles))
                user = {
                    'hub_id': current_user.hub_id,
                    'role': user_role['name'],
                    'note': user_form.note.data,
                    'birthday': user_form.birthday.data.isoformat() if user_form.birthday.data else None,
                }
                user_id = user_form.user_id.data
            else:
                user = {'role': current_user.role.name}
                user_id = current_user.id

            user['phone'] = user_form.about_user.phone.data
            user['location'] = user_form.about_user.location.data
            user['email_new'] = user_form.about_user.email_new.data
            user['email_modified'] = user_form.about_user.email_modified.data
            user['email_disapproved'] = user_form.about_user.email_disapproved.data
            user['email_approved'] = user_form.about_user.email_approved.data
            user['email_comment'] = user_form.about_user.email_comment.data
            user['name'] = user_form.about_user.full_name.data.strip()
            user['position'] = user_form.about_user.position.data.strip().lower()

            if user['role'] in ('purchaser', 'validator'):
                user['categories'] = user_form.about_user.categories.data
                user['projects'] = user_form.about_user.projects.data

            response = UserApi.put_entity(user_id, **user)

            if response is not None:
                flash('Данные успешно сохранены.')
            else:
                flash('Не удалось сохранить пользователя.')
        else:
            errors = (
                user_form.about_user.full_name.errors +
                user_form.about_user.phone.errors +
                user_form.about_user.categories.errors +
                user_form.about_user.projects.errors +
                user_form.about_user.position.errors +
                user_form.about_user.location.errors +
                user_form.about_user.email_new.errors +
                user_form.about_user.email_modified.errors +
                user_form.about_user.email_approved.errors +
                user_form.about_user.email_disapproved.errors
            )
            if isinstance(user_form, UserRolesForm):
                errors += (
                    user_form.user_id.errors +
                    user_form.role.errors +
                    user_form.note.errors +
                    user_form.birthday.errors
                )
            for error in errors:
                flash(error)
        return redirect(url_for('main.show_users'))


@bp.route('/users/remove/<int:user_id>', methods=['POST'])
@login_required
@role_required(['admin'])
def remove_user(user_id):
    response = UserApi.delete_entity(user_id)
    if response is None:
        flash('Не удалось удалить пользователя.')
    else:
        flash('Пользователь успешно удалён.')
    return redirect(url_for('main.show_users'))


@bp.route('/users/download')
@login_required
@role_required(['admin'])
def download_users():
    return '<p>download_users</p>'
    # users = User.query.filter(
    #     or_(User.role == UserRoles.default, User.hub_id == current_user.hub_id)
    # ).order_by(User.name, User.email).all()
    # wb = Workbook()
    # ws = wb.active
    # for i, header in enumerate([
    #     'Имя',
    #     'Телефон',
    #     'Email',
    #     'Роль',
    #     'Площадка',
    #     'Права',
    #     'Заметка',
    #     'Активность',
    #     'Регистрация',
    #     'Согласованных заявок пользователя',
    #     'Сумма согласованных заявок пользователя',
    #     'Согласовал заявок',
    #     'Должен согласовать заявок',
    #     'Номер для согласования',
    #     'День рождения'
    # ], start=1):
    #     ws.cell(1, i).value = header

    # for i, user in enumerate(users, start=2):
    #     ws.cell(i, 1).value = user.name
    #     ws.cell(i, 2).value = user.phone
    #     ws.cell(i, 3).value = user.email
    #     ws.cell(i, 4).value = user.position.name if user.position is not None else ''
    #     ws.cell(i, 5).value = user.location
    #     ws.cell(i, 6).value = user.role
    #     ws.cell(i, 7).value = user.note
    #     ws.cell(i, 8).value = user.last_seen
    #     ws.cell(i, 9).value = user.registered
    #     ws.cell(i, 15).value = user.birthday

    #     # Orders which user is initiative for
    #     orders = Order.query.filter_by(
    #         initiative_id=user.id,
    #         status=OrderStatus.approved
    #     ).all()
    #     ws.cell(i, 10).value = len(orders)
    #     ws.cell(i, 11).value = sum([o.total for o in orders])

    #     if user.role in [UserRoles.purchaser, UserRoles.validator]:
    #         # Orders approved by user
    #         orders = Order.query.filter_by(
    #             hub_id=current_user.hub_id
    #         ).join(OrderApproval).filter_by(
    #             user_id=user.id,
    #             product_id=None
    #         ).all()
    #         ws.cell(i, 12).value = len(orders)

    #         # Orders to be approved
    #         orders = Order.query.filter(
    #             Order.hub_id == current_user.hub_id,
    #             or_(
    #                 Order.status == OrderStatus.new,
    #                 Order.status == OrderStatus.partly_approved,
    #                 Order.status == OrderStatus.modified
    #             ),
    #             ~Order.user_approvals.any(OrderApproval.user_id == user.id),
    #             ~Order.children.any()
    #         )
    #         orders = orders.join(OrderPosition)
    #         orders = orders.filter_by(position_id=user.position_id)
    #         orders = orders.join(OrderCategory)
    #         orders = orders.filter(
    #             OrderCategory.category_id.in_([cat.id for cat in user.categories])
    #         )
    #         orders = orders.join(Site)
    #         orders = orders.filter(Site.project_id.in_([p.id for p in user.projects]))
    #         orders = orders.all()
    #         ws.cell(i, 13).value = len(orders)
    #         ws.cell(i, 14).value = ', '.join([o.number for o in orders])
    #     else:
    #         ws.cell(i, 12).value = 0
    #         ws.cell(i, 13).value = 0
    # data = save_virtual_workbook(wb)
    # return Response(
    #     data,
    #     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    #     headers={'Content-Disposition': 'attachment;filename=users.xlsx'}
    # )
